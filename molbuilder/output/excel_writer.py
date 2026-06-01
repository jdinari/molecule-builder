"""
excel_writer.py
===============
Write energetics results to a formatted Excel workbook.

Colour coding:
    Dark-blue header row
    Light-blue alternate data rows
    Yellow   -- structure did not converge
    Amber    -- bond STRETCHED (> 1.20x initial M-L length)
    Red      -- bond BROKEN    (> 1.35x initial M-L length, ligand likely dissociated)
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

_HDR_FILL   = "1F4E79"   # dark blue  -- header
_HDR_FONT   = "FFFFFF"   # white      -- header text
_ALT_FILL   = "D6E4F0"   # light blue -- alternate rows
_NOCONV_FILL= "FFF2CC"   # yellow     -- not converged
_STRETCH_FILL="FCE4D6"   # amber      -- bond stretched
_BROKEN_FILL = "FF0000"  # red        -- bond broken / dissociated


# -- column specification ------------------------------------------------------
# (header, row_key, col_width, number_format, centre)

_COLS = [
    ("Name",              "molecule_name",        32, None,        False),
    ("Formula",           "formula",              14, None,        True),
    ("Structure",         "structure",            20, None,        True),
    ("Ox state",          "ox_label",              8, None,        True),
    ("CN",                "cn",                    5, None,        True),
    ("Geometry",          "geometry",              8, None,        True),
    ("Charge",            "charge",                7, None,        True),
    ("Spin mult",         "spin_multiplicity",     9, None,        True),
    ("Arrangement",       "arrangement",          12, None,        True),
    ("Ligands",           "ligand_combo",         28, None,        False),
    ("# atoms",           "n_atoms",               7, None,        True),
    # -- xTB energetics --------------------------------------------------------
    ("E_xtb (eV)",        "relax_energy_eV",      13, "0.0000",    True),
    ("G_xtb (eV)",        "relax_gibbs_eV",       13, "0.0000",    True),
    ("ZPE_xtb (eV)",      "relax_zpe_eV",         12, "0.0000",    True),
    ("H_xtb (eV)",        "relax_enthalpy_eV",    12, "0.0000",    True),
    ("S_xtb (eV/K)",      "relax_entropy_eV_K",   13, "0.000000",  True),
    ("T (K)",             "relax_T_K",             7, "0.0",       True),
    ("P (Pa)",            "relax_P_Pa",            8, "0",         True),
    # -- MACE energetics -------------------------------------------------------
    ("E_mace (eV)",       "relax_mace_energy_eV", 13, "0.0000",    True),
    ("G_mace (eV)",       "relax_mace_gibbs_eV",  13, "0.0000",    True),
    ("DeltaE mace-xtb (eV)",  "relax_dE_mace_xtb_eV",13, "0.0000",    True),
    # -- quality / bond status -------------------------------------------------
    ("Converged",         "relax_converged",      10, None,        True),
    ("Steps",             "relax_steps",           7, None,        True),
    ("Backend",           "relax_backend",        10, None,        True),
    ("Bond status",       "bond_status",          12, None,        True),
    ("Max elongation",    "bond_max_elongation",  14, "0.000",     True),
    ("# broken bonds",    "bond_n_broken",        13, None,        True),
]


def write_energetics_excel(
    rows: List[Dict[str, Any]],
    path: str | Path,
    title: str = "Molbuilder energetics",
) -> Path:
    """
    Write energetics rows to a formatted .xlsx workbook.

    Parameters
    ----------
    rows  : Row dicts as returned by run_energetics().
    path  : Output file path.
    title : Title shown in the Summary sheet.

    Returns
    -------
    Path to the written .xlsx file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel output: pip install openpyxl"
        ) from exc

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb   = Workbook()

    # -- shared styles ---------------------------------------------------------
    def _font(bold=False, color="000000", size=10):
        return Font(name="Arial", bold=bold, color=color, size=size)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _border():
        s = Side(style="thin", color="AAAAAA")
        return Border(bottom=s)

    hdr_font   = _font(bold=True, color=_HDR_FONT)
    body_font  = _font()
    hdr_fill   = _fill(_HDR_FILL)
    alt_fill   = _fill(_ALT_FILL)
    noconv_fill= _fill(_NOCONV_FILL)
    stretch_fill=_fill(_STRETCH_FILL)
    broken_fill= _fill(_BROKEN_FILL)
    centre_al  = Alignment(horizontal="center", vertical="center")
    left_al    = Alignment(horizontal="left",   vertical="center")

    # -- Sheet 1: Energetics ---------------------------------------------------
    ws = wb.active
    ws.title         = "Energetics"
    ws.freeze_panes  = "A2"
    ws.row_dimensions[1].height = 22

    for ci, (hdr, _, width, _, _centre) in enumerate(_COLS, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = centre_al; c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    from molbuilder.energetics import BondStatus

    for ri, row in enumerate(rows, 2):
        bs   = row.get("bond_status", BondStatus.OK)
        conv = row.get("relax_converged", True)

        if bs == BondStatus.BROKEN:
            row_fill = broken_fill
        elif bs == BondStatus.STRETCHED:
            row_fill = stretch_fill
        elif conv is False:
            row_fill = noconv_fill
        elif ri % 2 == 0:
            row_fill = alt_fill
        else:
            row_fill = None

        for ci, (_, key, _, fmt, do_centre) in enumerate(_COLS, 1):
            val = row.get(key)
            if isinstance(val, bool):
                val = "Yes" if val else "No"
            elif val is None:
                val = ""
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = body_font
            c.alignment = centre_al if do_centre else left_al
            if fmt and val != "":
                c.number_format = fmt
            if row_fill:
                c.fill = row_fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLS))}1"

    # -- Sheet 2: Bond status legend -------------------------------------------
    ws_leg = wb.create_sheet("Legend")
    ws_leg.column_dimensions["A"].width = 16
    ws_leg.column_dimensions["B"].width = 55

    legend = [
        ("Bond status",   "Explanation",              True),
        (BondStatus.OK,       "All M-L bonds within 1.20x initial length",  False),
        (BondStatus.STRETCHED,"Longest M-L bond 1.20-1.35x initial (possible strain)", False),
        (BondStatus.BROKEN,   "At least one M-L bond > 1.35x initial (ligand likely dissociated)", False),
        ("", "", False),
        ("Row colour",    "Meaning",                  True),
        ("Blue (alt)",    "Clean structure",           False),
        ("Yellow",        "Geometry did not converge within step limit", False),
        ("Amber",         "Bond STRETCHED",            False),
        ("Red",           "Bond BROKEN -- review before DFT", False),
    ]
    fills = {
        "Blue (alt)": alt_fill, "Yellow": noconv_fill,
        "Amber": stretch_fill,  "Red": broken_fill,
    }
    for row_i, (a, b, bold) in enumerate(legend, 1):
        ca = ws_leg.cell(row_i, 1, a)
        cb = ws_leg.cell(row_i, 2, b)
        ca.font = cb.font = _font(bold=bold)
        if a in fills:
            ca.fill = cb.fill = fills[a]

    # -- Sheet 3: Summary ------------------------------------------------------
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = title
    ws2["A1"].font = _font(bold=True, size=13, color=_HDR_FILL)
    ws2.row_dimensions[1].height = 24
    for col, width in zip("ABCDE", [28, 16, 16, 16, 16]):
        ws2.column_dimensions[col].width = width

    from collections import Counter
    import statistics as _st

    by_type  = Counter(r.get("structure", "") for r in rows)
    by_ox    = Counter(r.get("ox_label",  "") for r in rows)
    by_bond  = Counter(r.get("bond_status", BondStatus.OK) for r in rows)

    def _sh(ws, row, col, val):
        c = ws.cell(row, col, val)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = centre_al
        return c
    def _sb(ws, row, col, val, fmt=None):
        c = ws.cell(row, col, val)
        c.font = body_font
        if fmt: c.number_format = fmt
        return c

    _sh(ws2, 3, 1, "Structure type"); _sh(ws2, 3, 2, "Count")
    for i, (k, v) in enumerate(sorted(by_type.items()), 4):
        _sb(ws2, i, 1, k); _sb(ws2, i, 2, v)

    r2 = 4 + len(by_type) + 2
    _sh(ws2, r2, 1, "Oxidation state"); _sh(ws2, r2, 2, "Count")
    for i, (k, v) in enumerate(sorted(by_ox.items()), r2+1):
        _sb(ws2, i, 1, k); _sb(ws2, i, 2, v)

    r3 = r2 + len(by_ox) + 3
    _sh(ws2, r3, 1, "Bond status"); _sh(ws2, r3, 2, "Count")
    bond_fills = {BondStatus.OK: alt_fill, BondStatus.STRETCHED: stretch_fill,
                  BondStatus.BROKEN: broken_fill}
    for i, (k, v) in enumerate(sorted(by_bond.items()), r3+1):
        ca = _sb(ws2, i, 1, k); cb = _sb(ws2, i, 2, v)
        f = bond_fills.get(k)
        if f:
            ca.fill = cb.fill = f

    e_vals = [r["relax_energy_eV"] for r in rows if r.get("relax_energy_eV") is not None]
    g_vals = [r["relax_gibbs_eV"]  for r in rows if r.get("relax_gibbs_eV")  is not None]
    if e_vals:
        r4 = r3 + len(by_bond) + 3
        _sh(ws2, r4,   1, "E_xtb (eV)"); _sh(ws2, r4, 2, "Min")
        _sh(ws2, r4, 3, "Max");          _sh(ws2, r4, 4, "Mean")
        for ci, v in zip([2,3,4], [min(e_vals), max(e_vals), _st.mean(e_vals)]):
            _sb(ws2, r4+1, ci, round(v, 4), "0.0000")
        if g_vals:
            _sh(ws2, r4+2, 1, "G_xtb (eV)")
            for ci, v in zip([2,3,4], [min(g_vals), max(g_vals), _st.mean(g_vals)]):
                _sb(ws2, r4+3, ci, round(v, 4), "0.0000")

    wb.save(path)
    return path
