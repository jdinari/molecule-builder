"""
tests/test_energetics.py
========================
Tests for molbuilder.energetics — the bond-status pipeline and run_energetics().
"""

import json
import numpy as np
import pytest

from molbuilder import build_isomers, dimer
from molbuilder.core.molecule import Molecule
from molbuilder.energetics import (
    run_energetics, molecule_name, BondStatus,
    _bond_status, _STRETCHED_THRESHOLD, _BROKEN_THRESHOLD,
)
from molbuilder.relaxation import check_bonds_intact


# ── fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def ni_h2o4():
    return build_isomers("Ni", ox=2, ligands=["H2O"] * 4, geometry="sqp")[0]


@pytest.fixture(scope="module")
def base_row(ni_h2o4):
    return {
        "filename":    "/tmp/test.POSCAR",
        "formula":     ni_h2o4.formula,
        "structure":   "monomer",
        "ox_label":    "NiII",
        "cn":          4,
        "geometry":    "sqp",
        "charge":      0,
        "n_atoms":     ni_h2o4.num_atoms(),
        "ligand_combo":"H2O4",
        "bridge":      "",
        "arrangement": "",
    }


# ── BondStatus constants ──────────────────────────────────────────────────────

class TestBondStatus:
    def test_constants(self):
        assert BondStatus.OK        == "OK"
        assert BondStatus.STRETCHED == "STRETCHED"
        assert BondStatus.BROKEN    == "BROKEN"

    def test_threshold_ok(self):
        assert _bond_status(1.00)  == BondStatus.OK
        assert _bond_status(1.10)  == BondStatus.OK
        assert _bond_status(_STRETCHED_THRESHOLD - 0.01) == BondStatus.OK

    def test_threshold_stretched(self):
        assert _bond_status(_STRETCHED_THRESHOLD + 0.01) == BondStatus.STRETCHED
        assert _bond_status(1.25) == BondStatus.STRETCHED
        assert _bond_status(_BROKEN_THRESHOLD - 0.01) == BondStatus.STRETCHED

    def test_threshold_broken(self):
        assert _bond_status(_BROKEN_THRESHOLD + 0.01) == BondStatus.BROKEN
        assert _bond_status(2.0)  == BondStatus.BROKEN
        assert _bond_status(3.5)  == BondStatus.BROKEN

    def test_exactly_at_stretched_boundary(self):
        # > 1.20 is stretched, == 1.20 is OK
        assert _bond_status(_STRETCHED_THRESHOLD)       == BondStatus.OK
        assert _bond_status(_STRETCHED_THRESHOLD + 1e-9)== BondStatus.STRETCHED

    def test_exactly_at_broken_boundary(self):
        assert _bond_status(_BROKEN_THRESHOLD)          == BondStatus.STRETCHED
        assert _bond_status(_BROKEN_THRESHOLD + 1e-9)   == BondStatus.BROKEN


# ── check_bonds_intact ────────────────────────────────────────────────────────

class TestCheckBondsIntact:
    def test_intact_structure(self, ni_h2o4):
        """Molecule compared to itself: all bonds intact, no elongation."""
        bc = check_bonds_intact(ni_h2o4, ni_h2o4)
        assert bc["intact"] is True
        assert bc["max_elongation"] == pytest.approx(1.0, abs=0.01)
        assert bc["broken_bonds"] == []

    def test_detects_broken_bond(self, ni_h2o4):
        """Moving an O 5 Å away should be flagged as broken."""
        mol2 = Molecule.from_json(ni_h2o4.to_json())
        for a in mol2.atoms:
            if a.symbol == "O":
                a.position = a.position + np.array([5., 0., 0.])
                break
        bc = check_bonds_intact(ni_h2o4, mol2)
        assert bc["intact"] is False
        assert bc["max_elongation"] > _BROKEN_THRESHOLD
        assert len(bc["broken_bonds"]) >= 1

    def test_broken_bond_has_required_keys(self, ni_h2o4):
        mol2 = Molecule.from_json(ni_h2o4.to_json())
        for a in mol2.atoms:
            if a.symbol == "O":
                a.position = a.position + np.array([5., 0., 0.])
                break
        bc = check_bonds_intact(ni_h2o4, mol2)
        bond = bc["broken_bonds"][0]
        for key in ("metal_idx", "ligand_idx", "d_before_A", "d_after_A", "elongation"):
            assert key in bond, f"Missing key: {key}"

    def test_elongation_is_ratio(self, ni_h2o4):
        mol2 = Molecule.from_json(ni_h2o4.to_json())
        for a in mol2.atoms:
            if a.symbol == "O":
                d_before = float(np.linalg.norm(a.position - ni_h2o4.atoms[0].position))
                a.position = a.position + np.array([5., 0., 0.])
                break
        bc = check_bonds_intact(ni_h2o4, mol2)
        bond = bc["broken_bonds"][0]
        expected_elong = bond["d_after_A"] / bond["d_before_A"]
        assert abs(bond["elongation"] - expected_elong) < 0.01

    def test_no_metal_symbol_returns_intact(self, ni_h2o4):
        """Molecule with no metal_symbol set should return intact=True (no bonds to check)."""
        mol_no_metal = Molecule.from_json(ni_h2o4.to_json())
        mol_no_metal.metal_symbol = ""
        bc = check_bonds_intact(mol_no_metal, mol_no_metal)
        assert bc["intact"] is True


# ── molecule_name ─────────────────────────────────────────────────────────────

class TestMoleculeName:
    def test_basic(self):
        row = {"ox_label": "NiII", "cn": 6, "geometry": "oct",
               "ligand_combo": "H2O6", "bridge": "", "arrangement": ""}
        n = molecule_name(row)
        assert "NiII" in n
        assert "CN6"  in n
        assert "oct"  in n

    def test_includes_bridge(self):
        row = {"ox_label": "NiII", "cn": 4, "geometry": "sqp",
               "ligand_combo": "H2O2", "bridge": "mu-OH", "arrangement": ""}
        n = molecule_name(row)
        assert "mu-OH" in n

    def test_includes_arrangement(self):
        row = {"ox_label": "NiIII", "cn": 5, "geometry": "sqpy",
               "ligand_combo": "bare", "bridge": "mu-HCOO",
               "arrangement": "triangular"}
        n = molecule_name(row)
        assert "triangular" in n

    def test_empty_row_doesnt_crash(self):
        n = molecule_name({})
        assert isinstance(n, str)


# ── constrain_bonds default ───────────────────────────────────────────────────

class TestConstrainBondsDefault:
    """constrain_bonds=False must be the default on all public functions."""
    import inspect

    def test_relax_default_false(self):
        import inspect
        from molbuilder import relax
        sig = inspect.signature(relax)
        assert sig.parameters["constrain_bonds"].default is False

    def test_thermochemistry_default_false(self):
        import inspect
        from molbuilder import thermochemistry
        sig = inspect.signature(thermochemistry)
        assert sig.parameters["constrain_bonds"].default is False

    def test_run_energetics_default_false(self):
        import inspect
        sig = inspect.signature(run_energetics)
        assert sig.parameters["constrain_bonds"].default is False


# ── run_energetics ────────────────────────────────────────────────────────────

class TestRunEnergetics:
    @pytest.fixture(scope="class")
    def result_row(self, ni_h2o4, base_row):
        updated = run_energetics(
            rows=[base_row],
            mols={base_row["filename"]: ni_h2o4},
            backend="xtb",
            compute_thermo=False,
            fmax=0.1,
            steps=30,
            constrain_bonds=False,
            verbose=False,
        )
        return updated[0]

    def test_returns_list(self, ni_h2o4, base_row):
        updated = run_energetics(
            rows=[base_row], mols={base_row["filename"]: ni_h2o4},
            backend="xtb", fmax=0.1, steps=30, verbose=False,
        )
        assert isinstance(updated, list)
        assert len(updated) == 1

    def test_molecule_name_set(self, result_row):
        assert result_row["molecule_name"] != ""
        assert "NiII" in result_row["molecule_name"]

    def test_energy_set(self, result_row):
        assert result_row["relax_energy_eV"] is not None
        assert isinstance(result_row["relax_energy_eV"], float)
        assert result_row["relax_energy_eV"] < 0

    def test_bond_status_set(self, result_row):
        assert result_row["bond_status"] in (
            BondStatus.OK, BondStatus.STRETCHED, BondStatus.BROKEN
        )

    def test_bond_max_elongation_set(self, result_row):
        assert result_row["bond_max_elongation"] is not None
        assert result_row["bond_max_elongation"] >= 1.0

    def test_bond_n_broken_set(self, result_row):
        assert result_row["bond_n_broken"] is not None
        assert isinstance(result_row["bond_n_broken"], int)
        assert result_row["bond_n_broken"] >= 0

    def test_spin_multiplicity_set(self, result_row, ni_h2o4):
        assert result_row["spin_multiplicity"] == ni_h2o4.spin_multiplicity

    def test_converged_is_python_bool(self, result_row):
        """Must be plain bool, not numpy.bool_ (JSON serialization safety)."""
        assert type(result_row["relax_converged"]) is bool

    def test_steps_is_python_int(self, result_row):
        assert type(result_row["relax_steps"]) is int

    def test_missing_mol_skipped_gracefully(self, base_row):
        updated = run_energetics(
            rows=[base_row], mols={},   # no molecule for this row
            backend="xtb", verbose=False,
        )
        assert len(updated) == 1
        assert updated[0]["relax_energy_eV"] is None

    def test_broken_bond_flagged(self, ni_h2o4, base_row):
        """Artificially broken molecule — bond should show strain or xTB may fail."""
        mol_broken = Molecule.from_json(ni_h2o4.to_json())
        for a in mol_broken.atoms:
            if a.symbol == "O":
                a.position = a.position + np.array([5., 0., 0.])
                break
        updated = run_energetics(
            rows=[base_row], mols={base_row["filename"]: mol_broken},
            backend="xtb", fmax=0.1, steps=5, verbose=False,
        )
        r = updated[0]
        # xTB may fail to converge on a severely broken structure (SCF non-convergence)
        # in which case bond_status is set to "ERROR". Both BROKEN/STRETCHED and ERROR
        # are acceptable outcomes — what we're testing is that the pipeline
        # doesn't silently report OK for a clearly broken structure.
        assert r["bond_status"] in (BondStatus.STRETCHED, BondStatus.BROKEN, "ERROR"), \
            f"Expected STRETCHED/BROKEN/ERROR for broken structure, got {r['bond_status']}"

    def test_csv_written(self, ni_h2o4, base_row, tmp_path):
        csv_path = tmp_path / "test.csv"
        run_energetics(
            rows=[base_row], mols={base_row["filename"]: ni_h2o4},
            backend="xtb", fmax=0.1, steps=20, verbose=False,
            csv_file=csv_path,
        )
        assert csv_path.exists()
        import csv
        with open(csv_path) as f:
            row = list(csv.DictReader(f))[0]
        assert "bond_status" in row
        assert "relax_energy_eV" in row

    def test_excel_written(self, ni_h2o4, base_row, tmp_path):
        xlsx_path = tmp_path / "test.xlsx"
        run_energetics(
            rows=[base_row], mols={base_row["filename"]: ni_h2o4},
            backend="xtb", fmax=0.1, steps=20, verbose=False,
            excel_file=xlsx_path,
        )
        assert xlsx_path.exists()
        assert xlsx_path.stat().st_size > 2000   # not an empty file

    def test_json_sidecar_written(self, ni_h2o4, base_row, tmp_path):
        import shutil
        poscar = tmp_path / "test.POSCAR"
        # Write a dummy POSCAR so the path exists
        poscar.write_text("dummy")
        row = {**base_row, "filename": str(poscar)}
        run_energetics(
            rows=[row], mols={str(poscar): ni_h2o4},
            backend="xtb", fmax=0.1, steps=20, verbose=False,
            output_dir=tmp_path,
        )
        # JSON sidecar should be written alongside the relaxed POSCAR
        json_files = list(tmp_path.glob("*.json"))
        assert len(json_files) >= 1
        data = json.loads(json_files[0].read_text())
        assert "bond_status" in data
        assert "energy_eV" in data
        assert isinstance(data.get("converged"), (bool, type(None)))


# ── Excel writer ──────────────────────────────────────────────────────────────

class TestWriteEnergeticsExcel:
    def _sample_rows(self):
        return [
            {"molecule_name": "NiII CN4 sqp H2O4", "formula": "H8NiO4",
             "structure": "monomer", "ox_label": "NiII", "cn": 4,
             "geometry": "sqp", "charge": 0, "spin_multiplicity": 3,
             "arrangement": "", "ligand_combo": "H2O4", "n_atoms": 13,
             "relax_energy_eV": -656.06, "relax_gibbs_eV": -654.58,
             "relax_zpe_eV": 1.91, "relax_enthalpy_eV": -653.71,
             "relax_entropy_eV_K": 0.00435, "relax_T_K": 298.15,
             "relax_P_Pa": 101325, "relax_converged": True, "relax_steps": 39,
             "relax_backend": "xtb", "relax_model": "GFN2-xTB",
             "relax_filename": "poscar/test.POSCAR",
             "relax_mace_energy_eV": None, "relax_mace_gibbs_eV": None,
             "relax_dE_mace_xtb_eV": None,
             "bond_status": BondStatus.OK, "bond_max_elongation": 1.05,
             "bond_n_broken": 0},
            {"molecule_name": "NiIII CN6 oct HCOO4_H2O2", "formula": "C4H8NiO10",
             "structure": "monomer", "ox_label": "NiIII", "cn": 6,
             "geometry": "oct", "charge": -1, "spin_multiplicity": 2,
             "arrangement": "", "ligand_combo": "HCOO4_H2O2", "n_atoms": 27,
             "relax_energy_eV": -800.12, "relax_gibbs_eV": None,
             "relax_zpe_eV": None, "relax_enthalpy_eV": None,
             "relax_entropy_eV_K": None, "relax_T_K": None, "relax_P_Pa": None,
             "relax_converged": False, "relax_steps": 300,
             "relax_backend": "xtb", "relax_model": "GFN2-xTB",
             "relax_filename": "poscar/test2.POSCAR",
             "relax_mace_energy_eV": None, "relax_mace_gibbs_eV": None,
             "relax_dE_mace_xtb_eV": None,
             "bond_status": BondStatus.BROKEN, "bond_max_elongation": 1.48,
             "bond_n_broken": 2},
        ]

    def test_creates_file(self, tmp_path):
        from molbuilder.output.excel_writer import write_energetics_excel
        p = write_energetics_excel(self._sample_rows(), tmp_path / "out.xlsx")
        assert p.exists()
        assert p.stat().st_size > 3000

    def test_three_sheets(self, tmp_path):
        from molbuilder.output.excel_writer import write_energetics_excel
        from openpyxl import load_workbook
        p = write_energetics_excel(self._sample_rows(), tmp_path / "out.xlsx")
        wb = load_workbook(p, data_only=True)
        assert "Energetics" in wb.sheetnames
        assert "Legend"     in wb.sheetnames
        assert "Summary"    in wb.sheetnames

    def test_data_rows_written(self, tmp_path):
        from molbuilder.output.excel_writer import write_energetics_excel
        from openpyxl import load_workbook
        rows = self._sample_rows()
        p = write_energetics_excel(rows, tmp_path / "out.xlsx")
        wb = load_workbook(p, data_only=True)
        ws = wb["Energetics"]
        # Row 1 = header, rows 2+ = data
        assert ws.max_row == len(rows) + 1

    def test_bond_status_column_present(self, tmp_path):
        from molbuilder.output.excel_writer import write_energetics_excel
        from openpyxl import load_workbook
        p = write_energetics_excel(self._sample_rows(), tmp_path / "out.xlsx")
        wb = load_workbook(p, data_only=True)
        ws = wb["Energetics"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "Bond status" in headers
